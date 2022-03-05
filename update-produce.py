#! python3
# update-produce.py - Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.get_highest_row()): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

# Loops over all the rows
# If the row is for garlic, celery, or lemons, changes the price

# Open the spreadsheet file.
# For each row, check whether the value in column A is Celery, Garlic or Lemon.
# If it is, update the price in column B.
# Save the spreadsheet to a new file (so that you don’t lose the old spread-sheet, just in case)